from typing import Optional, Tuple, Union
import customtkinter as ctk
from tkinter import *
from tkinter import messagebox
from tkinter import Button
import openpyxl, xlrd
import pathlib
from openpyxl import Workbook

#aparencia do sistema
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.layout_config()
        self.appearence()
        self.todo_sistema()
        self.frame()
    
    def layout_config(self):
        self.title("Gestão de Clientes")
        self.geometry("700x500")

    def appearence(self):
        self.lb_apm = ctk.CTkLabel(self, text="Tema", bg_color="transparent", text_color=['#000', "#fff"]).place(x=50, y=430)
        self.opt_apm = ctk.CTkOptionMenu(self, values=["Light", "Dark", "System"], command=self.change_apm).place(x=50, y=460)


    def todo_sistema(self):
        Frame = ctk.CTkFrame(self, width=700, height=50, corner_radius=0, bg_color="#00acee", fg_color="#00acee")
        Frame.place(x=0, y=10)
        title = ctk.CTkLabel(Frame, text="Sistema de Gestão de Clientes", font=("Century Gothic bold", 24), text_color="#fff", bg_color="#00acee").place(x=190, y=10)

        span = ctk.CTkLabel(self, text="Por Favor, preencher todos os campos do formulário!", font=("Century Gothic bold", 16), text_color=["#000", "#fff"]).place(x=50, y=70)

        excel = pathlib.Path("Clientes.xlsx")

        if excel.exists():
            pass
        else:
            excel=Workbook()
            folha=excel.active
            folha['A1']="Nome Completo"
            folha['B1']="Contato"
            folha['C1']="Idade"
            folha['D1']="Endereço"
            folha['E1']="Gênero"
            folha['F1']="Observações"

            excel.save("Clientes.xlsx")

        def submit():

            name = name_value.get()
            contato = contato_value.get()
            idade = idade_value.get()
            endereco = endereco_value.get()
            genero = genero_combobox.get()
            obs = obs_entry.get(0.0, END)

            if (name =="" or contato =="" or idade == "" or endereco == ""):
                messagebox.showerror("Sistema", "ERRO!! \nPor favor preencha todos os campos" )
            else:
                excel = openpyxl.load_workbook('Clientes.xlsx')
                folha = excel.active
                folha.cell(column=1, row=folha.max_row+1, value=name)
                folha.cell(column=2, row=folha.max_row, value=contato)
                folha.cell(column=3, row=folha.max_row, value=idade)
                folha.cell(column=4, row=folha.max_row, value=endereco)
                folha.cell(column=5, row=folha.max_row, value=genero)
                folha.cell(column=6, row=folha.max_row, value=obs)
                
            excel.save(r"Clientes.xlsx")
            messagebox.showinfo("Sistema", "Dados Salvos com sucesso!!")
        def clear():
            name_value.set("")
            contato_value.set("")
            idade_value.set("")
            endereco_value.set("")
            obs_entry.delete(0.0, END)

        #Variaveis
        name_value = StringVar()
        contato_value = StringVar()
        idade_value = StringVar()
        endereco_value = StringVar()

        #Entradas
        name_entry = ctk.CTkEntry(self, width=350, textvariable=name_value, font=("Century Gothic bold", 16), fg_color="transparent")
        contato_entry = ctk.CTkEntry(self, width=200, textvariable=contato_value, font=("Century Gothic bold", 16), fg_color="transparent")
        endereco_entry = ctk.CTkEntry(self, width=200, textvariable=endereco_value, font=("Century Gothic bold", 16), fg_color="transparent")
        idade_entry = ctk.CTkEntry(self, width=150, textvariable=idade_value, font=("Century Gothic bold", 16), fg_color="transparent")

        #ComboBox
        genero_combobox = ctk.CTkComboBox(self, values=["Masculino", "Feminino", "Outros", ""], font=("Numito", 14))
        genero_combobox.set("")

        #Entrada Textos OBS

        obs_entry = ctk.CTkTextbox(self, width=470, height=150, font=("arial", 18), border_color="#aaa", border_width=1.5, fg_color="transparent")

        #label
        lb_name = ctk.CTkLabel(self, text="Nome Completo:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_contato = ctk.CTkLabel(self, text="Contato:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_idade = ctk.CTkLabel(self, text="Idade:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_genero = ctk.CTkLabel(self, text="Gênero:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_endereco = ctk.CTkLabel(self, text="Endereço:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_obs = ctk.CTkLabel(self, text="Observações:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])

        btn_submit = ctk.CTkButton(self, text="Salvar Dados".upper(), command=submit, fg_color="#151", hover_color="#131").place(x=300, y=420)
        btn_clear = ctk.CTkButton(self, text="Limpar Dados".upper(), command=clear, fg_color="#555", hover_color="#333").place(x=500, y=420)

        

        #Posicionamento do OBJS
        
        lb_name.place(x=50, y=120)
        name_entry.place(x=50, y=150)

        lb_contato.place(x=450, y=120)
        contato_entry.place(x=450, y=150)

        lb_endereco.place(x=50, y=190)
        endereco_entry.place(x=50, y=220)

        lb_idade.place(x=300, y=190)
        idade_entry.place(x=300, y=220)

        lb_genero.place(x=500, y=190)
        genero_combobox.place(x=500, y=220)

        lb_obs.place(x=50, y=260)
        obs_entry.place(x=180, y=260)


    def change_apm(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)

if __name__=="__main__":
    app = App()
    app.mainloop()
